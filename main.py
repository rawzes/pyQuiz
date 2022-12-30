"""Module performs testing according to provided XLS"""
import sys
import random
import argparse
import openpyxl
from validators.validate import validate_answer
from messages.messages import *

menu_options = {
    1: 'Random terms',
    2: 'Random definitions',
    3: 'Exit',
}

tmp1 = []
tmp2 = []


def print_menu():
    for key, value_option in menu_options.items():
        print(f'{key} -- {value_option}')


def read_file(f_name):
    work_book = openpyxl.load_workbook(f_name)
    work_sheet = work_book.active

    # Iterate the loop to read the cell values
    for row in range(0, work_sheet.max_row):
        for col in work_sheet.iter_cols(0, 1):
            tmp1.append(col[row].value)

    for row in range(0, work_sheet.max_row):
        for col in work_sheet.iter_cols(2, 2):
            tmp2.append(col[row].value)


def ask_question(terms, index):
    print('-' * 100)
    print(f'What is: {terms[index]}?')


def get_answer():
    answer = ''
    while answer == '':
        answer = input('Answer: ')
        if answer == '':
            print(EMPTY_ANSWER_ERROR)
    return answer


def ask_and_check(mode):
    terms = [i for i in tmp1 if i is not None]
    definitions = [i for i in tmp2 if i is not None]
    size = len(terms) - 1
    answer = ''
    while answer != 'x':
        random_index = random.randint(0, size)
        if mode == 1:
            ask_question(terms, random_index)
            answer = get_answer()
            if validate_answer(answer, definitions[random_index]):
                if answer in terms[random_index]:
                    print(SUCCESS)
                    print()
                else:
                    print(ERROR)
                    print('Правильно вот так: ' + definitions[random_index])
                    print()
        elif mode == 2:
            ask_question(definitions, random_index)
            answer = get_answer()
            if validate_answer(answer, definitions[random_index]):
                if answer in terms[random_index]:
                    print(SUCCESS)
                    print()
                else:
                    print(ERROR)
                    print('Правильно вот так: ' + terms[random_index])
                    print()


def start_test(mode, name):
    print('Нажми X для выхода')
    read_file(name)
    ask_and_check(mode)


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--name', type=str, required=True)
    args = parser.parse_args()
    while True:
        print_menu()
        OPTION = ''
        try:
            OPTION = int(input('Enter your choice: '))
        except ValueError:
            print('Wrong input. Please enter a number ...')

        if OPTION == 1:
            print(args.name)
            start_test(1, args.name)
        elif OPTION == 2:
            start_test(2, args.name)
        elif OPTION == 3:
            print('Thanks you for using...!')
            sys.exit()
        else:
            print('Invalid option. Please enter a number between 1 and 3.')
